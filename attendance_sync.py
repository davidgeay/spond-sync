# attendance_sync.py
# Secrets: SPOND_USERNAME, SPOND_PASSWORD, GOOGLE_SERVICE_ACCOUNT_JSON, SHEET_ID
# Deps:    spond gspread google-auth pandas openpyxl python-dateutil

import os, io, json, asyncio, re
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

import pandas as pd
from dateutil import parser as dtparser
import gspread
from google.oauth2.service_account import Credentials
from spond import spond

# ---------------- Config ----------------
GROUP_NAME = "IHKS G2008b/G2009b"

TIMEZONE = ZoneInfo(os.getenv("TIMEZONE", "Europe/Oslo"))
CUTOFF_LOCAL = datetime(2025, 8, 1, 0, 0, tzinfo=TIMEZONE)
CUTOFF_UTC = CUTOFF_LOCAL.astimezone(timezone.utc)
NOW_UTC = datetime.now(timezone.utc)

SHEET_ATT = "Attendance"
SHEET_DBG = "Debug"
FIXED_COLS = ["Player", "Total Present", "Total Missed", "Total Unanswered"]

ISTR_PAT = re.compile(r"\bistrening\b", re.IGNORECASE)

def log(msg: str): print(f"[spond-sync] {msg}", flush=True)

# ---------------- Sheets helpers ----------------
def sheets_client() -> gspread.Client:
    svc = json.loads(os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"])
    creds = Credentials.from_service_account_info(
        svc, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    return gspread.authorize(creds)

def open_spreadsheet(gc: gspread.Client):
    return gc.open_by_key(os.environ["SHEET_ID"])

def get_or_create_ws(sh, title: str):
    try:
        return sh.worksheet(title)
    except gspread.WorksheetNotFound:
        return sh.add_worksheet(title=title, rows=2000, cols=26)
    except gspread.exceptions.APIError as e:
        if "already exists" in str(e).lower():
            return sh.worksheet(title)
        raise

# ---------------- Generic utils ----------------
def _pick(d: Dict[str, Any], *keys):
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
    return None

def to_text(v: Any) -> str:
    if v is None: return ""
    if isinstance(v, (str, int, float, bool)): return str(v)
    if isinstance(v, dict):
        return " | ".join([f"{k}: {to_text(x)}" for k, x in v.items() if x is not None])
    if isinstance(v, (list, tuple, set)):
        return " | ".join([to_text(x) for x in v if x is not None])
    return str(v)

def contains_istrening(*vals: Any) -> bool:
    for v in vals:
        if ISTR_PAT.search(to_text(v)): return True
    return False

def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())

def parse_start_utc(d: Dict[str, Any]) -> Optional[datetime]:
    raw = _pick(d, "startTimeUtc", "start_time_utc", "startTime", "start",
                "startAt", "start_at", "startDateTime", "start_datetime",
                "utcStart", "utc_start", "startTimestamp", "start_timestamp")
    if raw is None: return None
    try:
        if isinstance(raw, (int, float)) and raw > 1_000_000_000:
            return datetime.fromtimestamp(float(raw), tz=timezone.utc)
        s = str(raw).replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None

def parse_datetime_from_text(text: str) -> Optional[datetime]:
    try:
        dt = dtparser.parse(text, fuzzy=True, dayfirst=True)
        if dt.tzinfo is None: dt = dt.replace(tzinfo=TIMEZONE)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None

# ---------------- Attendance parsing ----------------
# Role filtering: include only PLAYERS, exclude admins/coaches/tutors/guardians/parents etc.
ROLE_INCLUDE_WORDS = {
    "player","spiller","deltaker","member","utøver","athlete","keeper","målvakt"
}
ROLE_EXCLUDE_WORDS = {
    "leader","leder","coach","trener","admin","administrator",
    "lagleder","oppmann","team leader","staff",
    "tutor","guardian","parent","foresatt","forelder"
}

# Column candidates (EN + NO)
NAME_CANDS   = ["Name","Navn","Member name","Member","Participant","Deltaker","Spiller","Player"]
STATUS_CANDS = ["Status","Response","Svar","Svarstatus","Attendance","Attending","RSVP"]
REASON_CANDS = ["Note","Reason","Absence reason","Kommentar","Begrunnelse","Fraværsgrunn","Årsak","Notes","Message"]
ROLE_CANDS   = ["Type","Role","Rolle","Kategori","Category","Group","Gruppe"]

YES_HDR_HINTS    = re.compile(r"(attend|kommer|deltar|påmeld|present|going|ja|møter)", re.I)
NO_HDR_HINTS     = re.compile(r"(declin|kommer\s*ikke|deltar\s*ikke|nei|absent|not\s*going|kan\s*ikke)", re.I)
NORES_HDR_HINTS  = re.compile(r"(no\s*resp|ubesvar|ingen\s*svar|not\s*respond)", re.I)

def _pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = [str(c).strip() for c in df.columns]
    for c in candidates:
        if c in cols: return c
    low = [c.lower() for c in cols]
    for cand in candidates:
        cl = cand.lower()
        for i, l in enumerate(low):
            if cl in l: return cols[i]
    return None

def _truthy(v: Any) -> bool:
    s = str(v).strip().lower()
    return s in {"1","true","yes","ja","x","✓","present","checked","check","kommer","deltar","påmeldt","going"}

def infer_status_from_row(row: pd.Series, name_col: str, status_col: Optional[str]) -> Tuple[str, str]:
    reason = ""
    if status_col:
        raw = str(row.get(status_col, "")).strip()
        if raw:
            return raw, reason
    # boolean columns?
    for col in row.index:
        if col == name_col: continue
        val = row[col]
        if YES_HDR_HINTS.search(str(col)) and _truthy(val):
            return "Present", ""
        if NO_HDR_HINTS.search(str(col)) and _truthy(val):
            return "Absent", str(val)
        if NORES_HDR_HINTS.search(str(col)) and _truthy(val):
            return "No response", ""
    # scan whole row text
    blob = " ".join([str(v) for k, v in row.items() if k != name_col and pd.notna(v)]).lower()
    if re.search(r"(accepted|attending|present|going|kommer|deltar|påmeldt|ja|møter)", blob):
        return "Present",""
    if re.search(r"(declin|kommer\s*ikke|deltar\s*ikke|nei|fravær|kan\s*ikke|not\s*going)", blob):
        return "Absent",""
    if re.search(r"(no\s*response|ubesvart|ingen\s*svar|not\s*respond)", blob):
        return "No response",""
    return "",""

def xlsx_all_text(xbytes: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xbytes), data_only=True)
        parts = []
        for ws in wb.worksheets:
            for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 200), values_only=True):
                for c in r:
                    if isinstance(c, str) and c.strip():
                        parts.append(c.strip())
        return "\n".join(parts)
    except Exception:
        return ""

def read_attendance_xlsx(xbytes: bytes) -> Tuple[pd.DataFrame, Dict[str,str]]:
    """
    Returns:
      table: Member | Raw Status | Raw Reason
      roles_by_name: normalized_name -> role_text (best effort)
    """
    try:
        sheets = pd.read_excel(io.BytesIO(xbytes), sheet_name=None)
    except Exception:
        return pd.DataFrame(columns=["Member","Raw Status","Raw Reason"]), {}

    frames: List[pd.DataFrame] = []
    roles: Dict[str,str] = {}

    for _, df in (sheets or {}).items():
        if df is None or df.empty: 
            continue
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]

        name_col   = _pick_col(df, NAME_CANDS)
        status_col = _pick_col(df, STATUS_CANDS)
        reason_col = _pick_col(df, REASON_CANDS)
        role_col   = _pick_col(df, ROLE_CANDS)

        if not name_col:
            continue

        members, statuses, reasons = [], [], []
        for _, r in df.iterrows():
            name = str(r.get(name_col, "")).strip()
            if not name: 
                continue

            # capture role text (if any)
            role_text = str(r.get(role_col, "")).strip() if role_col else ""
            if role_text:
                roles.setdefault(normalize_name(name), role_text)

            raw_status, raw_reason = infer_status_from_row(r, name_col, status_col)
            if not raw_status and status_col:
                raw_status = str(r.get(status_col, "")).strip()
            if not raw_reason and reason_col:
                raw_reason = str(r.get(reason_col, "")).strip()

            members.append(name)
            statuses.append(raw_status)
            reasons.append(raw_reason)

        if members:
            frames.append(pd.DataFrame({"Member": members, "Raw Status": statuses, "Raw Reason": reasons}))

    if not frames:
        return pd.DataFrame(columns=["Member","Raw Status","Raw Reason"]), roles

    return pd.concat(frames, ignore_index=True).fillna(""), roles

# ---------------- JSON attendee extraction ----------------
def normalize_status(raw: Any) -> str:
    if raw is None: return "No response"
    s = str(raw).strip().lower()

    tokens_yes = ("yes","accepted","attending","present","going","kommer","deltar","påmeldt","ja","møter","checked_in","checkedin","checked-in")
    tokens_no  = ("no ","declin","absent","kommer ikke","deltar ikke","nei","fravær","kan ikke","not going","rejected")
    tokens_nr  = ("no response","unknown","maybe","usikker","kanskje","ubesvart","ingen svar","pending","not responded","har ikke svart")

    if any(t in s for t in tokens_yes): return "Present"
    if any(t in s for t in tokens_no):  return "Absent"
    if any(t in s for t in tokens_nr):  return "No response"
    if s in {"","-","—","nan"}: return "No response"
    return "No response"

def extract_from_json(d: Dict[str,Any]) -> Tuple[Dict[str,Tuple[str,str,str]], Dict[str,str]]:
    """
    Returns:
      statuses_by_name: norm_name -> (display_name, status, role_text)
      roles_by_name:    norm_name -> role_text
    """
    by_name: Dict[str, Tuple[str,str,str]] = {}
    roles: Dict[str,str] = {}

    arrays = []
    for k in ("participants","members","invites","responses","attendances"):
        v = d.get(k)
        if isinstance(v, list): arrays.append(v)

    for arr in arrays:
        for p in arr:
            name = _pick(p, "name","fullName","memberName","title","displayName") or ""
            name = str(name).strip()
            if not name: 
                continue
            nkey = normalize_name(name)

            role = str(_pick(p, "role","type","kategori","category","memberType") or "").strip().lower()
            if role: roles.setdefault(nkey, role)

            st = _pick(p, "status","response","rsvpStatus","attendanceStatus")
            if st is None:
                # booleans?
                if str(p.get("isAttending", "")).lower() in {"true","1"}: st = "attending"
                elif str(p.get("declined","")).lower() in {"true","1"}:   st = "declined"
                else: st = ""
            norm = normalize_status(st)
            by_name[nkey] = (name, norm, role)

    return by_name, roles

# ---------------- Status to cell ----------------
def status_to_cell(status: str, reason: str) -> Tuple[str, str]:
    st = normalize_status(status)
    if st == "Present":
        return ("Present", "present")
    if st == "Absent":
        txt = "Absent" + (f" — {str(reason).strip()}" if str(reason).strip() else "")
        return (txt, "absent")
    return ("No response", "noresp")

# ---------------- Spond flows ----------------
async def resolve_group_id(sp: spond.Spond) -> Optional[str]:
    groups = await sp.get_groups()
    log(f"Found {len(groups)} groups.")
    gid = None
    for g in groups:
        nm = _pick(g, "name","title","groupName") or ""
        gid_candidate = _pick(g, "id","groupId","uid")
        log(f"  - {gid_candidate} | {nm}")
        if nm.strip().lower() == GROUP_NAME.strip().lower():
            gid = gid_candidate
    if gid:
        log(f"Using group: {GROUP_NAME} (id={gid})")
    else:
        log(f"ERROR: Group '{GROUP_NAME}' not found.")
    return gid

def xlsx_contains_istrening(xbytes: bytes) -> bool:
    return contains_istrening(xlsx_all_text(xbytes))

# ---------------- Collect ----------------
async def collect_events_and_attendance() -> Tuple[List[str], Dict[str, Dict[str, Tuple[str, str]]], List[Dict[str, Any]], Dict[str,bool]]:
    sp = spond.Spond(username=os.environ["SPOND_USERNAME"], password=os.environ["SPOND_PASSWORD"])
    try:
        gid = await resolve_group_id(sp)
        if not gid:
            return [], {}, [], {}

        log(f"Fetching events (UTC {CUTOFF_UTC.isoformat()} → {NOW_UTC.isoformat()}) ...")
        try:
            events = await sp.get_events(group_id=gid, min_start=CUTOFF_UTC, max_start=NOW_UTC,
                                         include_scheduled=True, max_events=500)
        except TypeError:
            events = await sp.get_events(min_start=CUTOFF_UTC, max_start=NOW_UTC)

        log(f"Fetched {len(events)} events (list may be minimal).")

        included: List[Tuple[str, str, Optional[datetime]]] = []
        dbg: List[Dict[str, Any]] = []

        for le in events:
            eid = _pick(le, "id","eventId","uid")
            if not eid: 
                continue

            # Full event
            try:
                ev = await sp.get_event(eid)
            except Exception as e:
                log(f"WARNING get_event {eid}: {e}")
                continue

            title = _pick(ev, "title","name","eventName","subject") or ""
            start_utc = parse_start_utc(ev)
            start_disp = start_utc.isoformat() if start_utc else "NO-START"

            matched = ""
            if contains_istrening(ev):
                matched = "details"
            else:
                try:
                    xb = await sp.get_event_attendance_xlsx(eid)
                    if xlsx_contains_istrening(xb):
                        matched = "xlsx"
                    if not start_utc:
                        guess = parse_datetime_from_text(xlsx_all_text(xb))
                        if guess: start_utc = guess
                except Exception:
                    pass

            date_ok = True if start_utc is None else (start_utc >= CUTOFF_UTC)
            include = (matched != "") and date_ok

            dbg.append({
                "Event ID": eid,
                "Event Title": title[:200],
                "Start UTC": start_disp,
                "Matched (details/xlsx)": matched or "no",
                "Cutoff Date OK": "Yes" if date_ok else "No",
                "Included": "Yes" if include else "No",
            })

            if not include:
                continue

            header = (f"{start_utc.astimezone(TIMEZONE):%Y-%m-%d %H:%M} — {title or '(no title)'}"
                      if start_utc else f"(no time) — {title or '(no title)'}")
            included.append((eid, header, start_utc))

        included.sort(key=lambda t: (t[2] or CUTOFF_UTC))
        event_headers = [h for _, h, _ in included]

        per_member: Dict[str, Dict[str, Tuple[str, str]]] = {}
        global_role_flags: Dict[str, bool] = {}  # norm_name -> True if looks like player

        for (eid, header, _) in included:
            # JSON attendees
            json_map: Dict[str, Tuple[str,str,str]] = {}
            json_roles: Dict[str,str] = {}
            try:
                ev = await sp.get_event(eid)
                json_map, json_roles = extract_from_json(ev)
            except Exception:
                pass

            # XLSX attendees
            table = pd.DataFrame()
            x_roles: Dict[str,str] = {}
            try:
                xb = await sp.get_event_attendance_xlsx(eid)
                table, x_roles = read_attendance_xlsx(xb)
            except Exception:
                pass

            # Merge role info for relevance filter (players only)
            role_by_name: Dict[str,str] = {}
            role_by_name.update(json_roles)
            role_by_name.update(x_roles)

            def is_player_role(role_text: str) -> bool:
                t = (role_text or "").lower()
                if any(w in t for w in ROLE_EXCLUDE_WORDS): return False
                if any(w in t for w in ROLE_INCLUDE_WORDS): return True
                # if unknown, leave undecided (we’ll infer from name; default include later)
                return True

            # Build statuses for this event
            # 1) Use JSON if present, else XLSX, else No response
            used_names: set = set()

            for nkey, (disp_name, st, role_text) in json_map.items():
                txt, code = status_to_cell(st, "")
                per_member.setdefault(disp_name, {})[header] = (txt, code)
                used_names.add(normalize_name(disp_name))
                global_role_flags.setdefault(nkey, is_player_role(role_text))

            if not table.empty:
                for _, row in table.iterrows():
                    name = str(row.get("Member","")).strip()
                    if not name: 
                        continue
                    nkey = normalize_name(name)
                    # capture relevance from XLSX role if we haven’t seen the name
                    if nkey not in global_role_flags:
                        role_guess = ""
                        if nkey in x_roles: role_guess = x_roles[nkey]
                        global_role_flags[nkey] = is_player_role(role_guess)

                    if nkey in used_names:
                        continue  # JSON already provided a status
                    raw_st = str(row.get("Raw Status","")).strip()
                    raw_re = str(row.get("Raw Reason","")).strip()
                    txt, code = status_to_cell(raw_st, raw_re)
                    per_member.setdefault(name, {})[header] = (txt, code)

        # Keep only relevant members (players)
        # Build a map display_name -> relevance by looking up normalized name
        relevant_map: Dict[str,bool] = {}
        for disp in list(per_member.keys()):
            nkey = normalize_name(disp)
            relevant = global_role_flags.get(nkey, True)  # default True if unknown
            # extra safety: exclude if name contains leader/coach/admin etc
            nm = disp.lower()
            if any(w in nm for w in ROLE_EXCLUDE_WORDS):
                relevant = False
            relevant_map[disp] = relevant

        # prune non-players
        for disp in list(per_member.keys()):
            if not relevant_map.get(disp, True):
                per_member.pop(disp, None)

        return event_headers, per_member, dbg, relevant_map

    finally:
        try:
            if sp and getattr(sp, "clientsession", None):
                await sp.clientsession.close()
        except Exception:
            pass

# ---------------- Build sheet ----------------
def build_matrix(event_headers: List[str], per_member: Dict[str, Dict[str, Tuple[str, str]]]) -> Tuple[List[List[str]], List[List[str]]]:
    members = sorted(per_member.keys(), key=lambda s: s.lower())

    values: List[List[str]] = []
    colors: List[List[str]] = []

    for m in members:
        row_vals = [m, "", "", ""]
        row_cols: List[str] = []
        present = absent = noresp = 0

        for eh in event_headers:
            text, code = per_member.get(m, {}).get(eh, ("No response","noresp"))
            row_vals.append(text)
            row_cols.append(code)
            if code == "present": present += 1
            elif code == "absent": absent += 1
            else: noresp += 1

        missed = absent + noresp
        row_vals[1] = str(present)
        row_vals[2] = str(missed)
        row_vals[3] = str(noresp)

        values.append(row_vals)
        colors.append(row_cols)

    return values, colors

def write_wide_sheet(sh, ws_att, event_headers: List[str], values: List[List[str]], colors: List[List[str]]):
    header = FIXED_COLS + event_headers
    ws_att.clear()
    ws_att.update([header] + values)

    # Freeze header + name column
    try:
        sh.batch_update({
            "requests": [
                {"updateSheetProperties": {
                    "properties": {"sheetId": ws_att.id, "gridProperties": {"frozenRowCount": 1, "frozenColumnCount": 1}},
                    "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"
                }}
            ]
        })
    except Exception:
        pass

    color_map = {
        "present": {"red": 0.85, "green": 0.95, "blue": 0.85},
        "absent":  {"red": 0.98, "green": 0.85, "blue": 0.85},
        "noresp":  {"red": 0.93, "green": 0.86, "blue": 0.97},
    }

    requests = []
    n_rows, n_events = len(values), len(event_headers)
    if n_rows > 0 and n_events > 0:
        for r in range(n_rows):
            for c in range(n_events):
                code = colors[r][c]
                rgb = color_map.get(code)
                if not rgb: continue
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_att.id,
                            "startRowIndex": 1 + r,
                            "endRowIndex":   2 + r,
                            "startColumnIndex": 4 + c,
                            "endColumnIndex":   5 + c,
                        },
                        "cell": {"userEnteredFormat": {"backgroundColor": rgb}},
                        "fields": "userEnteredFormat.backgroundColor"
                    }
                })
        for i in range(0, len(requests), 5000):
            sh.batch_update({"requests": requests[i:i+5000]})

    log("Attendance sheet updated (players only).")

def write_debug(ws_dbg, dbg_rows: List[Dict[str, Any]], relevance: Dict[str,bool]):
    cols = ["Event ID","Event Title","Start UTC","Matched (details/xlsx)","Cutoff Date OK","Included"]
    ws_dbg.clear()
    data = [cols] + [[r.get(c,"") for c in cols] for r in dbg_rows] if dbg_rows else [cols, ["(no events)", "", "", "", "", ""]]
    ws_dbg.update(data)

    # small appendix showing who was filtered out as non-player
    if relevance:
        keep = sorted([n for n, ok in relevance.items() if ok])
        drop = sorted([n for n, ok in relevance.items() if not ok])
        ws_dbg.append_row([])
        ws_dbg.append_row(["Kept (players):"] + keep[:20])
        ws_dbg.append_row(["Excluded (non-players):"] + drop[:20])

    log("Debug sheet updated.")

# ---------------- Main ----------------
async def main():
    gc = sheets_client()
    sh = open_spreadsheet(gc)
    ws_att = get_or_create_ws(sh, SHEET_ATT)
    ws_dbg = get_or_create_ws(sh, SHEET_DBG)

    headers, per_member, dbg, relevance = await collect_events_and_attendance()
    values, colors = build_matrix(headers, per_member)

    write_wide_sheet(sh, ws_att, headers, values, colors)
    write_debug(get_or_create_ws(sh, SHEET_DBG), dbg, relevance)
    log("Sync complete.")

if __name__ == "__main__":
    asyncio.run(main())
